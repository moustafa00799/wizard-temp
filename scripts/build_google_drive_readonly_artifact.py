from __future__ import annotations

import hashlib
import json
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT = Path(os.environ.get('CDKS_GOOGLE_DRIVE_NORMALIZED_ROOT', '.local/private-research/google-drive/2026-08-27'))
EXPORTS = Path(os.environ.get('CDKS_GOOGLE_DRIVE_EXPORT_MANIFEST', str(ROOT / 'export-manifest.jsonl')))
OUT = Path(os.environ.get('CDKS_GOOGLE_DRIVE_NORMALIZED_OUTPUT', str(ROOT / 'normalized-drive-evidence-2026-08-27.json')))

FORBIDDEN_TEXT = re.compile(r'(password|passwd|token|cookie|secret|api.?key|access.?key|refresh.?token|iban|credit.?card|phone|mobile|email|address|full.?name|customer|client|الاسم|الهاتف|البريد|العنوان|كلمة السر|رمز)', re.I)
DATE_PATTERN = re.compile(r'(?<!\d)(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})(?!\d)')
CURRENCY_PATTERN = re.compile(r'\b(EGP|SAR|USD|AED|EUR|جنيه|ريال|دولار)\b', re.I)

TOKEN_MAP: list[tuple[str, re.Pattern[str]]] = [
    ('active_users', re.compile(r'(active users?|users?|المستخدمون? النشطون?|مستخدمون? نشطون?)', re.I)),
    ('new_users', re.compile(r'(new users?|المستخدمون? الجدد|المستخدمين? الجدد)', re.I)),
    ('sessions', re.compile(r'(sessions?|الجلسات)', re.I)),
    ('engaged_sessions', re.compile(r'(engaged sessions?|الجلسات المتفاعلة)', re.I)),
    ('engagement_rate', re.compile(r'(engagement rate|نسبة المشاركة|معدل المشاركة|التفاعل)', re.I)),
    ('engagement_duration', re.compile(r'(engagement duration|average engagement|مدة التفاعل)', re.I)),
    ('events', re.compile(r'(event count|events?|عدد الأحداث|الأحداث)', re.I)),
    ('key_events', re.compile(r'(key events?|conversions?|الأحداث الرئيسية|الإحالات الناجحة)', re.I)),
    ('views', re.compile(r'(views?|screen views?|المشاهدات)', re.I)),
    ('revenue', re.compile(r'(revenue|earnings|profits?|الأرباح|الإيرادات)', re.I)),
    ('ad_spend', re.compile(r'(ad spend|advertising cost|cost of ads|تكلفة الإعلانات)', re.I)),
    ('ad_clicks', re.compile(r'(ad clicks?|clicks? on ads|النقرات على الإعلانات)', re.I)),
    ('cpc', re.compile(r'(cost per click|cpc|تكلفة النقرة)', re.I)),
    ('ctr', re.compile(r'(click.?through rate|ctr|نسبة النقر)', re.I)),
    ('impressions', re.compile(r'(impressions?|مرات الظهور)', re.I)),
    ('position', re.compile(r'(average position|position|متوسط موضع)', re.I)),
    ('average_monthly_searches', re.compile(r'(avg.? monthly searches|average monthly searches|عمليات البحث الشهرية)', re.I)),
    ('competition', re.compile(r'(competition|المنافسة)', re.I)),
    ('bid', re.compile(r'(bid|top of page|عرض السعر)', re.I)),
    ('price', re.compile(r'(^|[^a-z])(price|sale.?price|السعر)([^a-z]|$)', re.I)),
    ('quantity', re.compile(r'(quantity|الكمية)', re.I)),
    ('budget', re.compile(r'(budget|الميزانية)', re.I)),
    ('target_achievement', re.compile(r'(target|ach|achievement|الهدف|الإنجاز)', re.I)),
]
DIMENSION_MAP: list[tuple[str, re.Pattern[str]]] = [
    ('date', re.compile(r'(date|day|week|month|year|التاريخ|اليوم|الأسبوع|الشهر|السنة)', re.I)),
    ('channel', re.compile(r'(channel|source|medium|القنوات|المصدر|الوسيط)', re.I)),
    ('campaign', re.compile(r'(campaign|الحملة|الحملات)', re.I)),
    ('event', re.compile(r'(event name|event|اسم الحدث|نوع الحدث)', re.I)),
    ('device', re.compile(r'(device|browser|operating system|الجهاز|المتصفح|نظام التشغيل)', re.I)),
    ('geography', re.compile(r'(country|city|location|البلد|الدولة|المدينة|الموقع)', re.I)),
    ('audience', re.compile(r'(audience|cohort|segment|الجمهور|الشريحة|المجموعة النموذجية)', re.I)),
    ('page_value_omitted', re.compile(r'(page|landing|screen|url|path|الصفحة|المقصودة|الشاشة|الرابط)', re.I)),
    ('query_value_omitted', re.compile(r'(query|keyword|search term|سلسلة طلب البحث|كلمة مفتاحية|عبارة البحث)', re.I)),
    ('product_value_omitted', re.compile(r'(product|sku|title|brand|المنتج|معرف المنتج|العنوان|العلامة)', re.I)),
]


def normalize_digits(value: str) -> str:
    trans = str.maketrans('٠١٢٣٤٥٦٧٨٩٫٬', '0123456789.,')
    return value.translate(trans)


def as_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if str(value) not in {'nan', 'inf', '-inf'} else None
    text = normalize_digits(str(value)).strip().replace(',', '')
    if not text or text in {'-', '—', 'n/a', 'NA', 'null'}:
        return None
    if text.endswith('%'):
        text = text[:-1].strip()
    try:
        number = float(text)
        return number if number == number and abs(number) != float('inf') else None
    except ValueError:
        return None


def date_token(value: Any) -> str | None:
    if isinstance(value, datetime):
        if not (1900 <= value.year <= 2100):
            return None
        return value.date().isoformat()
    text = normalize_digits(str(value)) if value is not None else ''
    match = DATE_PATTERN.search(text)
    if not match:
        return None
    year, month, day = map(int, match.groups())
    if not (1900 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31):
        return None
    return f'{year:04d}-{month:02d}-{day:02d}'


def token_set(headers: Iterable[str], mapping: list[tuple[str, re.Pattern[str]]]) -> list[str]:
    tokens: set[str] = set()
    for header in headers:
        for token, pattern in mapping:
            if pattern.search(header):
                tokens.add(token)
    return sorted(tokens)


def compact_schema(rows: list[tuple[Any, ...]]) -> tuple[int, list[str], list[int]]:
    if not rows:
        return 0, [], []
    scan = rows[:25]
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(scan):
        values = [str(value).strip() for value in row if value not in (None, '')]
        if not values:
            continue
        score = len(token_set(values, TOKEN_MAP)) + len(token_set(values, DIMENSION_MAP))
        score += sum(1 for value in values if re.search(r'[A-Za-zء-ي]', value))
        if score > best_score:
            best_idx, best_score = idx, score
    headers = [str(value).strip() for value in rows[best_idx] if value not in (None, '')]
    indices = [idx for idx, value in enumerate(rows[best_idx]) if value not in (None, '')]
    return best_idx, headers, indices


def safe_metric_aggregates(rows: list[tuple[Any, ...]], header_idx: int, headers: list[str], indices: list[int]) -> tuple[dict[str, Any], list[str], list[str]]:
    metric_by_col: dict[int, str] = {}
    dimension_tokens = token_set(headers, DIMENSION_MAP)
    for header, idx in zip(headers, indices):
        matches = [token for token, pattern in TOKEN_MAP if pattern.search(header)]
        if matches:
            metric_by_col[idx] = matches[0]
    aggregates: dict[str, dict[str, float | int | None]] = {}
    date_values: list[str] = []
    pii_signal = False
    for row in rows[header_idx + 1:]:
        row_text = ' '.join(str(value) for value in row if value not in (None, ''))
        if FORBIDDEN_TEXT.search(row_text):
            pii_signal = True
        for idx, metric in metric_by_col.items():
            if idx >= len(row):
                continue
            number = as_number(row[idx])
            if number is None:
                date_value = date_token(row[idx])
                if date_value:
                    date_values.append(date_value)
                continue
            current = aggregates.setdefault(metric, {'numericCellCount': 0, 'sum': 0.0, 'min': number, 'max': number})
            current['numericCellCount'] = int(current['numericCellCount']) + 1
            current['sum'] = float(current['sum']) + number
            current['min'] = min(float(current['min']), number)
            current['max'] = max(float(current['max']), number)
        for value in row:
            value_date = date_token(value)
            if value_date:
                date_values.append(value_date)
    normalized = {}
    for metric, stats in aggregates.items():
        normalized[metric] = {
            'numericCellCount': int(stats['numericCellCount']),
            'sum': round(float(stats['sum']), 8),
            'min': round(float(stats['min']), 8),
            'max': round(float(stats['max']), 8),
        }
    return normalized, sorted(set(date_values)), dimension_tokens + (['pii_signal_detected_omitted'] if pii_signal else [])


def file_record(meta: dict[str, Any]) -> dict[str, Any]:
    path = Path(meta['rawPath'])
    result: dict[str, Any] = {
        'sourceRef': 'drive-file-sha256:' + meta['sha256'][:16],
        'label': meta['label'],
        'dataClass': meta['dataClass'],
        'rawSha256': meta['sha256'],
        'rawSizeBytes': int(meta['sizeBytes']),
        'rawRowsOmitted': True,
        'rawValuesOmitted': True,
        'rows': [],
        'sheets': [],
        'metricAvailability': {},
        'period': None,
        'dimensions': [],
        'flags': [],
        'scope': {'market': None, 'industry': None, 'locale': None, 'currency': None, 'verified': False},
    }
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        result['flags'] = ['workbook_read_error']
        result['errorClass'] = type(exc).__name__
        return result
    all_metrics: dict[str, dict[str, float | int | None]] = {}
    all_dates: list[str] = []
    all_dims: set[str] = set()
    any_pii_signal = False
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        header_idx, headers, indices = compact_schema(rows)
        metrics, dates, dims = safe_metric_aggregates(rows, header_idx, headers, indices)
        any_pii_signal = any_pii_signal or 'pii_signal_detected_omitted' in dims
        all_dates.extend(dates)
        all_dims.update(dims)
        for metric, stats in metrics.items():
            current = all_metrics.setdefault(metric, {'numericCellCount': 0, 'sum': 0.0, 'min': stats['min'], 'max': stats['max']})
            current['numericCellCount'] = int(current['numericCellCount']) + int(stats['numericCellCount'])
            current['sum'] = round(float(current['sum']) + float(stats['sum']), 8)
            current['min'] = min(float(current['min']), float(stats['min']))
            current['max'] = max(float(current['max']), float(stats['max']))
        result['sheets'].append({'titleHash': hashlib.sha256(worksheet.title.encode('utf-8')).hexdigest()[:16], 'rowCount': max(0, len(rows) - (header_idx + 1 if rows else 0)), 'columnCount': len(headers), 'metricTokens': sorted(metrics), 'dimensionTokens': sorted(set(dims) - {'pii_signal_detected_omitted'})})
    result['metricAvailability'] = {metric: {'numericCellCount': int(stats['numericCellCount']), 'sum': round(float(stats['sum']), 8), 'min': round(float(stats['min']), 8), 'max': round(float(stats['max']), 8)} for metric, stats in sorted(all_metrics.items())}
    result['dimensions'] = sorted(all_dims - {'pii_signal_detected_omitted'})
    if any_pii_signal:
        result['flags'].append('pii_signal_detected_omitted')
    if all_dates:
        result['period'] = {'min': min(all_dates), 'max': max(all_dates)}
    if result['dataClass'] in {'ga4', 'ga4_ads_linked'}:
        result['scope'] = {'market': 'SA', 'industry': 'interior_design_and_decoration', 'locale': 'ar', 'currency': None, 'verified': False, 'verificationNote': 'User-confirmed activity/market; Drive property, currency, and exact site identity were not proven in this pass.'}
    elif result['dataClass'] == 'search_console':
        result['scope'] = {'market': 'SA', 'industry': 'interior_design_and_decoration', 'locale': 'ar', 'currency': None, 'verified': False, 'verificationNote': 'User-confirmed Saudi interior-design activity; Search Console property/site ownership and exact scope require verification.'}
    elif result['dataClass'] == 'keyword_planner':
        result['scope'] = {'market': None, 'industry': None, 'locale': None, 'currency': None, 'verified': False, 'verificationNote': 'Keyword Planner scope, location, language, method, and date were not established from sanitized structural parsing.'}
    elif result['dataClass'] in {'store_product', 'catalog_feed'}:
        result['flags'].append('unverified_catalog_candidate')
        result['scope']['verificationNote'] = 'Catalog identity, ownership, market, currency, and Easy Orders linkage are unverified.'
    elif result['dataClass'] in {'campaign_report', 'sales_report', 'seller_profile'}:
        result['flags'].append('scope_unverified')
    return result


def structural_fingerprint(record: dict[str, Any]) -> str:
    stable = {
        'dataClass': record['dataClass'],
        'sheets': record['sheets'],
        'metricAvailability': sorted(record['metricAvailability']),
        'period': record['period'],
    }
    return hashlib.sha256(json.dumps(stable, sort_keys=True, ensure_ascii=False).encode('utf-8')).hexdigest()


def main() -> None:
    metas = [json.loads(line) for line in EXPORTS.read_text(encoding='utf-8').splitlines() if line.strip()]
    records = [file_record(meta) for meta in metas]
    seen: dict[tuple[str, str, str], int] = {}
    for index, record in enumerate(records):
        fingerprint = structural_fingerprint(record)
        record['structuralFingerprint'] = fingerprint
        period = record.get('period') or {}
        duplicate_key: tuple[str, str, str] | None = None
        if record['dataClass'] == 'search_console' and period.get('min') == '2024-01-01' and period.get('max') == '2024-12-25':
            duplicate_key = ('search_console_known_pair', fingerprint, f"{period.get('min')}:{period.get('max')}")
        elif record['dataClass'] == 'catalog_feed' and record.get('rawSizeBytes') == 72002:
            duplicate_key = ('catalog_exact_copy', fingerprint, str(record.get('rawSizeBytes')))
        if duplicate_key is not None and duplicate_key in seen:
            record['duplicateOfIndex'] = seen[duplicate_key]
            record['flags'].append('known_duplicate_excluded')
        elif duplicate_key is not None:
            seen[duplicate_key] = index
    # Do not expose local paths, Drive IDs, names, URLs, values, or headers in the normalized artifact.
    output = {
        'contractVersion': '1.0',
        'generatedAt': '2026-08-27T00:00:00.000Z',
        'provider': 'google_drive_readonly',
        'authorizationScope': 'user_provided_drive_export_read_only',
        'recordCount': len(records),
        'records': records,
        'policy': {
            'readOnlyDrive': True,
            'driveWrites': False,
            'rawRowsOmitted': True,
            'rawValuesOmitted': True,
            'freeTextOmitted': True,
            'urlsOmitted': True,
            'queriesAndKeywordsOmitted': True,
            'customerIdentityOmitted': True,
            'credentialsOmitted': True,
            'marketValidated': False,
            'canonicalBlueprintMutation': False,
            'easyOrdersAttachment': False,
        },
    }
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    counts = Counter(record['dataClass'] for record in records)
    flags = Counter(flag for record in records for flag in record['flags'])
    print(json.dumps({'status': 'PASS', 'recordCount': len(records), 'classCounts': dict(sorted(counts.items())), 'flagCounts': dict(sorted(flags.items())), 'duplicateCount': sum('structural_duplicate_excluded' in record['flags'] for record in records), 'output': str(OUT)}, ensure_ascii=False))


if __name__ == '__main__':
    main()
